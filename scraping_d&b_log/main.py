import os
import re
import csv
import datetime
import time
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

# --- 設定項目 ---
OUTPUT_EXCEL_PATH = Path.home() / "Desktop" / "dnb_eventlog.xlsx"
HEADLESS = False  # デバッグ時は False に設定
DEFAULT_WAIT_TIME = 5000  # ms

def parse_ip_range(ip_str):
    """
    "192.168.13.10-20,25,30-32" のような形式をパースして IP リストを返す
    """
    ips = []
    # 共通のサブネット部分を抽出 (簡易的な実装)
    match = re.match(r"(\d+\.\d+\.\d+\.)", ip_str)
    if not match:
        return [ip_str] if ip_str else []
    
    base_ip = match.group(1)
    parts = ip_str.replace(base_ip, "").split(",")
    
    for part in parts:
        if "-" in part:
            start, end = map(int, part.split("-"))
            for i in range(start, end + 1):
                ips.append(f"{base_ip}{i}")
        else:
            ips.append(f"{base_ip}{part}")
    return ips

def download_csv(page):
    """
    Event Log ページで件数を選択し、CSV をダウンロードする
    """
    # 件数ドロップダウンの選択肢 (多い順に試行)
    counts = ["1000", "500", "200", "100"]
    
    for count in counts:
        try:
            print(f"  - 取得件数 {count} を試行中...")
            
            # TODO: ドロップダウンのセレクタを実機に合わせて修正してください
            # 例: page.select_option("select#log-count", value=count)
            # ここでは仮の操作
            dropdown_selector = "select.event-log-count" # 仮
            if page.is_visible(dropdown_selector, timeout=2000):
                page.select_option(dropdown_selector, value=count)
                page.wait_for_timeout(1000)

            # CSVダウンロードボタンのクリックと待機
            # TODO: ダウンロードボタンのセレクタを実機に合わせて修正してください
            download_button_selector = "button#download-csv" # 仮
            
            with page.expect_download(timeout=10000) as download_info:
                page.click(download_button_selector)
            
            download = download_info.value
            temp_path = Path(f"./temp_{int(time.time())}.csv")
            download.save_as(temp_path)
            
            print(f"  - CSV ダウンロード成功 ({count}件)")
            return temp_path, count
            
        except Exception as e:
            print(f"  - {count}件での取得に失敗しました: {e}")
            continue
            
    return None, None

def collect_event_log(ip):
    """
    1台のアンプからログを取得する
    """
    result = {
        "ip": ip,
        "status": "INIT",
        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "model": "Unknown",
        "csv_path": None,
        "count": 0,
        "error": None
    }
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()
        
        try:
            print(f"--- IP: {ip} 接続開始 ---")
            url = f"http://{ip}"
            page.goto(url, timeout=15000)
            
            # TODO: アンプのモデル名（D20/D40）を取得するロジック
            # 例: result["model"] = page.inner_text(".device-model")
            result["model"] = "D20" # 仮
            
            # TODO: Event Log ページへの遷移ボタンをクリック
            # page.click("text=Event Log")
            # page.wait_for_load_state("networkidle")
            
            csv_path, count = download_csv(page)
            
            if csv_path:
                result["status"] = "OK"
                result["csv_path"] = csv_path
                result["count"] = count
            else:
                result["status"] = "CSV取得失敗"
                
        except PlaywrightTimeoutError:
            result["status"] = "TIMEOUT"
            print(f"  - タイムアウトしました")
        except Exception as e:
            result["status"] = "ERROR"
            result["error"] = str(e)
            print(f"  - エラー発生: {e}")
        finally:
            browser.close()
            
    return result

def write_amp_sheet(wb, amp_result):
    """
    個別アンプのシートを作成・更新する
    """
    if amp_result["status"] != "OK":
        return

    sheet_name = f"ID_{amp_result['ip'].split('.')[-1]}_{amp_result['model']}"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(title=sheet_name)
    
    # 基本情報
    ws.append(["取得日時", amp_result["timestamp"]])
    ws.append(["IPアドレス", amp_result["ip"]])
    ws.append(["取得件数", amp_result["count"]])
    ws.append([]) # 空行
    
    # CSVの内容を書き込み
    try:
        df = pd.read_csv(amp_result["csv_path"])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 簡易的なテーブル設定（フィルタ有効化）
        ws.auto_filter.ref = f"A5:{chr(64 + len(df.columns))}{ws.max_row}" # noqa
        
        # 一時ファイルの削除
        if os.path.exists(amp_result["csv_path"]):
            os.remove(amp_result["csv_path"])
            
    except Exception as e:
        print(f"  - Excel書き込みエラー: {e}")

def write_summary_sheet(wb, results):
    """
    SUMMARY シートを作成する
    """
    if "SUMMARY" in wb.sheetnames:
        del wb["SUMMARY"]
    
    ws = wb.create_sheet(title="SUMMARY", index=0)
    headers = ["IPアドレス", "状態", "モデル", "取得件数", "取得日時", "エラー内容"]
    ws.append(headers)
    
    for res in results:
        ws.append([
            res["ip"],
            res["status"],
            res["model"],
            res["count"],
            res["timestamp"],
            res.get("error", "")
        ])

def main():
    print("=== d&b Event Log Collector ===")
    ip_input = input("IP範囲を入力してください (例: 192.168.13.10-20,25): ")
    if not ip_input:
        print("IPが入力されませんでした。終了します。")
        return

    target_ips = parse_ip_range(ip_input)
    print(f"ターゲットIP: {target_ips}")

    results = []
    
    # Excelブックの準備
    if OUTPUT_EXCEL_PATH.exists():
        wb = load_workbook(OUTPUT_EXCEL_PATH)
    else:
        wb = Workbook()
        # デフォルトのシートを削除
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for ip in target_ips:
        res = collect_event_log(ip)
        results.append(res)
        
        if res["status"] == "OK":
            write_amp_sheet(wb, res)
        
        print(f"--- IP: {ip} 完了 (Status: {res['status']}) ---")

    write_summary_sheet(wb, results)
    
    wb.save(OUTPUT_EXCEL_PATH)
    print(f"\n全処理完了。出力先: {OUTPUT_EXCEL_PATH}")

if __name__ == "__main__":
    main()
