import sys
import webbrowser
from threading import Timer, Thread
import signal
from flask import Flask, render_template, request, session, jsonify, send_file
import sqlite3
from pathlib import Path
import io
import openpyxl
import unicodedata
import shutil
import tempfile
import os
import csv
from datetime import datetime, timedelta
import webview
import socket
import time
import base64
import logging

# --- ログ設定 (macOS標準の場所: ~/Library/Logs/RF_Unyo_System/ ) ---
def setup_logging():
    log_dir = Path.home() / "Library" / "Logs" / "RF_Unyo_System"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / "debug.log"

    # --- ログファイル内の古い行(3か月以上前)を削除する ---
    if log_file.exists():
        try:
            cutoff_date = datetime.now() - timedelta(days=90)
            cutoff_str = cutoff_date.strftime('%Y-%m-%d')
            
            new_lines = []
            with open(log_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            keep_this_entry = True
            for line in lines:
                if len(line) >= 10 and line[0:4].isdigit() and line[4] == '-' and line[7] == '-':
                    if line[0:10] >= cutoff_str: keep_this_entry = True
                    else: keep_this_entry = False
                
                if keep_this_entry: new_lines.append(line)
            
            with open(log_file, 'w', encoding='utf-8') as f:
                f.writelines(new_lines)
        except Exception as e:
            print(f"Failed to clean up log entries: {e}")

    # ロギングの初期化
    logging.basicConfig(
        filename=str(log_file),
        level=logging.INFO,
        format='%(asctime)s %(levelname)s: %(message)s',
        encoding='utf-8'
    )
    
    # Flask(Werkzeug)のアクセスログを抑制（重要ログのみ表示）
    werkzeug_log = logging.getLogger('werkzeug')
    werkzeug_log.setLevel(logging.ERROR)
    
    logging.info("--- Application Started ---")
    return True

setup_logging()

# --- 保存用APIクラス ---
class Api:
    def __init__(self):
        self.window = None

    def save_file(self, data_base64, default_filename):
        try:
            # 最新の書き方 (webview.FileDialog.SAVE) に修正
            file_path = self.window.create_file_dialog(
                webview.FileDialog.SAVE, 
                directory=str(Path.home() / "Desktop"), 
                save_filename=default_filename
            )
            
            if file_path:
                if isinstance(file_path, (list, tuple)):
                    if not file_path: return False
                    file_path = file_path[0]
                with open(file_path, 'wb') as f:
                    f.write(base64.b64decode(data_base64))
                logging.info(f"File saved: {file_path}")
                return True
        except Exception as e:
            logging.error(f"Save file error: {e}")
        return False

    def export_log(self):
        try:
            src = Path.home() / "Library" / "Logs" / "RF_Unyo_System" / "debug.log"
            dst = Path.home() / "Desktop" / f"rf_unyo_debug_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            if src.exists():
                shutil.copy2(src, dst)
                logging.info(f"Log exported to Desktop: {dst.name}")
                return {"status": "success", "filename": dst.name}
            else:
                return {"status": "error", "message": "ログファイルが見つかりません。"}
        except Exception as e:
            logging.error(f"Log export error: {e}")
            return {"status": "error", "message": str(e)}

api = Api()

# --- パス解決の設定 ---
def get_base_path():
    if hasattr(sys, '_MEIPASS'): return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent

BASE_DIR = get_base_path()
DB_PATH = BASE_DIR / "database.db"
MASTER_XLSX = BASE_DIR / "masters" / "master.xlsx"

app = Flask(__name__, 
            template_folder=str(BASE_DIR / "templates"),
            static_folder=str(BASE_DIR / "static"))
app.secret_key = "rf_unyo_secret_key"

# --- バージョン設定 ---
APP_VERSION = "1.0.0"
DATA_VERSION = "20250831"
VERSION_STR = f"v{APP_VERSION}_{DATA_VERSION}"

@app.context_processor
def inject_version():
    return {'version_str': VERSION_STR}

def normalize_text(text):
    if text is None: return ""
    return unicodedata.normalize('NFKC', str(text)).lower()

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.create_function("NORM", 1, normalize_text)
    return conn

@app.route("/")
def index():
    if "keep_list" not in session: session["keep_list"] = []
    return render_template("index.html")

@app.route("/adjustment")
def adjustment():
    try:
        logging.info("Navigating to /adjustment")
        keep_list = session.get("keep_list", [])
        conn = get_db_connection()
        devices = conn.execute("SELECT * FROM devices").fetchall()
        tv_channels = conn.execute("SELECT * FROM tv_channels").fetchall()
        conn.close()
        dev_list = [dict(d) for d in devices]
        ch_list = [dict(c) for c in tv_channels]
        for d in ch_list:
            if 'TVchannel' in d:
                try: d['TVchannel'] = int(d['TVchannel'])
                except: pass
        return render_template("adjustment.html", venues=keep_list, devices=dev_list, tv_channels=ch_list)
    except Exception as e:
        logging.error(f"Error in /adjustment: {e}")
        return str(e), 500

@app.route("/settings")
def settings():
    return render_template("settings.html")

@app.route("/get_settings")
def get_settings():
    try:
        conn = get_db_connection()
        m = conn.execute("SELECT * FROM member_info WHERE id=1").fetchone()
        u = conn.execute("SELECT * FROM onsite_user WHERE id=1").fetchone()
        conn.close()
        return jsonify({"member": dict(m) if m else {}, "user": dict(u) if u else {}})
    except Exception as e:
        logging.error(f"Error in /get_settings: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/save_settings", methods=["POST"])
def save_settings():
    try:
        d = request.json; m, u = d.get("member", {}), d.get("user", {})
        conn = get_db_connection()
        conn.execute("UPDATE member_info SET member_num1=?, member_num2=?, member_name=?, department=?, manager=?, tel=?, email=? WHERE id=1",
                    (m.get("member_num1"), m.get("member_num2"), m.get("member_name"), m.get("department"), m.get("manager"), m.get("tel"), m.get("email")))
        conn.execute("UPDATE onsite_user SET name=?, furigana=?, tel=?, email=? WHERE id=1",
                    (u.get("name"), u.get("furigana"), u.get("tel"), u.get("email")))
        conn.commit(); conn.close()
        return jsonify({"status": "success"})
    except Exception as e:
        logging.error(f"Error in /save_settings: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/search")
def search():
    try:
        query = request.args.get("q", "")
        if not query: return jsonify([])
        conn = get_db_connection()
        sql = "SELECT * FROM venues WHERE NORM(施設名) LIKE ? OR NORM(住所) LIKE ? OR NORM(都道府県名) LIKE ? LIMIT 100"
        results = conn.execute(sql, (f"%{normalize_text(query)}%",)*3).fetchall()
        conn.close()
        return jsonify([dict(row) for row in results])
    except Exception as e:
        logging.error(f"Error in /search: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/keep", methods=["POST"])
def keep():
    data = request.json.get("data")
    keep_list = session.get("keep_list", [])
    if not any(v["施設名"] == data["施設名"] and v["住所"] == data["住所"] for v in keep_list):
        keep_list.append(data); session["keep_list"] = keep_list; session.modified = True
    return jsonify({"status": "success", "count": len(keep_list)})

@app.route("/unkeep", methods=["POST"])
def unkeep():
    data = request.json.get("data")
    keep_list = [v for v in session.get("keep_list", []) if not (v["施設名"] == data["施設名"] and v["住所"] == data["住所"])]
    session["keep_list"] = keep_list; session.modified = True
    return jsonify({"status": "success", "count": len(keep_list)})

@app.route("/get_keep_list")
def get_keep_list(): return jsonify(session.get("keep_list", []))

@app.route("/export", methods=["POST"])
def export():
    try:
        data = request.json.get("data", [])
        conn = get_db_connection()
        member = conn.execute("SELECT * FROM member_info WHERE id = 1").fetchone()
        onsite = conn.execute("SELECT * FROM onsite_user WHERE id = 1").fetchone()
        conn.close()
        temp_dir = tempfile.mkdtemp(); temp_path = os.path.join(temp_dir, "temp.xlsx")
        shutil.copy2(MASTER_XLSX, temp_path)
        wb = openpyxl.load_workbook(temp_path)
        SHEETS = ["master_01", "master_02", "master_03"]
        for sn in SHEETS:
            if sn not in wb.sheetnames: continue
            ws = wb[sn]; ws.cell(row=4, column=4, value="新規")
            if member:
                ws.cell(row=4, column=13, value=member["member_num1"]); ws.cell(row=4, column=16, value=member["member_num2"]); ws.cell(row=4, column=23, value=member["member_name"])
                ws.cell(row=6, column=13, value=member["department"]); ws.cell(row=6, column=23, value=member["manager"]); ws.cell(row=8, column=13, value=member["tel"]); ws.cell(row=8, column=23, value=member["email"])
            if onsite:
                ws.cell(row=13, column=15, value=onsite["name"] + (f"（{onsite['furigana']}）" if onsite["furigana"] else ""))
                ws.cell(row=15, column=12, value=onsite["tel"]); ws.cell(row=15, column=23, value=onsite["email"])
        for i, item in enumerate(data[:12]):
            ws = wb[SHEETS[i // 4]]; row = 36 + (i % 4 * 12); v, chs = item["venue"], item["selected_channels"]
            ws.cell(row=row, column=10, value=v.get("郵便番号", "")); ws.cell(row=row, column=18, value=f"{v.get('都道府県名', '')}{v.get('住所', '')}")
            ws.cell(row=row+2, column=12, value=v.get("屋内外", "")); ws.cell(row=row+2, column=18, value=v.get("施設名", ""))
            ws.cell(row=row+4, column=15, value=v.get("適用エリア", "")); ws.cell(row=row+6, column=15, value=", ".join([str(c) for c in chs]))
        wb.save(temp_path)
        output = io.BytesIO()
        with open(temp_path, "rb") as f: output.write(f.read())
        output.seek(0); shutil.rmtree(temp_dir)
        logging.info("Excel export completed successfully.")
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        logging.error(f"Export error: {e}")
        return str(e), 500

@app.route("/export_wsm", methods=["POST"])
def export_wsm():
    try:
        data = request.json; venue = data.get("venue"); selected_channels = data.get("selected_channels", [])
        conn = get_db_connection()
        tv_ch_map = {r["TVchannel"]: r for r in conn.execute("SELECT * FROM tv_channels").fetchall()}
        conn.close()
        output = io.StringIO(); writer = csv.writer(output, lineterminator='\n')
        writer.writerow(["name", "type", "frequency", "tolerance", "minfrequency", "maxfrequency", "priority", "squelchlevel"])
        for ch in range(13, 54):
            is_available = venue.get(f"{ch}CH") == '○'; is_selected = ch in selected_channels; r = tv_ch_map.get(ch)
            if not r: continue
            min_f, max_f = r["minfrequency"], r["maxfrequency"]
            if ch > 13 and is_available != (venue.get(f"{ch-1}CH") == '○'):
                if is_available: min_f += 1000
                else: min_f -= 1000
            if ch < 53 and is_available != (venue.get(f"{ch+1}CH") == '○'):
                if is_available: max_f -= 1000
                else: max_f += 1000
            writer.writerow([f"TV {ch}", 2 if is_selected else 3, 0, 0, min_f, max_f, 2 if is_selected else 4, 5])
        writer.writerow(["Blocked", 3, 0, 0, 714000, 798000, 4, 5])
        mem = io.BytesIO(); mem.write(output.getvalue().encode('utf-8')); mem.seek(0)
        logging.info(f"WSM CSV export completed for: {venue.get('施設名')}")
        return send_file(mem, mimetype="text/csv")
    except Exception as e:
        logging.error(f"WSM Export error: {e}")
        return str(e), 500

@app.route("/shutdown", methods=["POST"])
def shutdown():
    def kill_server(): os.kill(os.getpid(), signal.SIGTERM)
    Timer(1.0, kill_server).start(); return jsonify({"status": "success"})

def run_flask():
    app.run(host="127.0.0.1", port=5001, debug=False, use_reloader=False)

def wait_for_server():
    for _ in range(50):
        try:
            with socket.create_connection(("127.0.0.1", 5001), timeout=1): return True
        except: time.sleep(0.1)
    return False

if __name__ == "__main__":
    t = Thread(target=run_flask); t.daemon = True; t.start()
    if wait_for_server():
        window = webview.create_window("RFチャンネルリスト検索システム", "http://127.0.0.1:5001", js_api=api, width=1200, height=800)
        api.window = window
        webview.start(debug=False)
    os.kill(os.getpid(), signal.SIGTERM)
